import re
import json
import unicodedata
import pandas as pd
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PASTA_BASE = Path(r"C:\ML_Analytics")
PASTA_RELATORIOS = PASTA_BASE / "relatorios_ml"
PASTA_SAIDA = PASTA_BASE / "saida_skus"
PASTA_JSON = PASTA_BASE / "saida_json"
ARQUIVO_HISTORICO = PASTA_BASE / "historico_geral.json"
ARQUIVO_MAPEAMENTO = PASTA_BASE / "mapeamento.xlsx"

PASTA_SAIDA.mkdir(exist_ok=True)
PASTA_JSON.mkdir(exist_ok=True)

COLUNAS_FINAIS = [
    "Data Análise",
    "Valores análise período anterior",
    "Ref.",
    "Valor Produto",
    "Vendas Brutas",
    "Comparado c/ o dia anterior Vendas Brutas",
    "Visitas",
    "Comparado c/ o dia anterior Visitas",
    "Conversão %",
    "Quantidade de Vendas",
    "% de Participação",
    "Qualidade do Anúncio",
    "Experiência de Compra"
]


def extrair_data(nome):
    m = re.search(r"(\d{4})_(\d{2})_(\d{2})", nome)
    if not m:
        return None
    return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3))).date()


def normalizar_texto(txt):
    txt = str(txt or "").strip().lower()
    txt = unicodedata.normalize("NFD", txt)
    txt = "".join(c for c in txt if unicodedata.category(c) != "Mn")
    txt = txt.replace("\n", " ")
    txt = " ".join(txt.split())
    return txt


def achar_coluna(df, possibilidades):
    colunas_norm = {normalizar_texto(col): col for col in df.columns}

    for possivel in possibilidades:
        p = normalizar_texto(possivel)
        for col_norm, col_original in colunas_norm.items():
            if p in col_norm:
                return col_original

    return None


def valor_linha(row, coluna, padrao=0):
    if coluna is None:
        return padrao
    return row.get(coluna, padrao)


def limpar_numero(v):
    if pd.isna(v) or v == "":
        return 0

    if isinstance(v, (int, float)):
        return float(v)

    s = str(v).strip().replace("R$", "").replace(" ", "")

    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(".", "")

    try:
        return float(s)
    except Exception:
        return 0


def limpar_inteiro_ml(v):
    if pd.isna(v) or v == "":
        return 0

    s = str(v).strip().replace(" ", "")

    if "." in s and "," not in s:
        partes = s.split(".")
        if len(partes[-1]) == 3:
            s = s.replace(".", "")
            return int(float(s))

    if "," in s:
        s = s.replace(".", "").replace(",", ".")

    try:
        n = float(s)

        if 0 < n < 10 and not n.is_integer():
            return int(round(n * 1000))

        return int(round(n))
    except Exception:
        return 0


def limpar_pct(v):
    if pd.isna(v) or v == "":
        return 0

    if isinstance(v, (int, float)):
        return float(v) if v <= 1 else float(v) / 100

    s = str(v).strip().replace("%", "").replace(" ", "")

    if "," in s:
        s = s.replace(".", "").replace(",", ".")

    try:
        return float(s) / 100
    except Exception:
        return 0


def br_moeda(valor):
    if valor == "-" or pd.isna(valor):
        return "-"

    return (
        f"R$ {float(valor):,.2f}"
        .replace(",", "X")
        .replace(".", ",")
        .replace("X", ".")
    )


def br_pct(valor):
    if valor == "-" or pd.isna(valor):
        return "-"

    return f"{float(valor) * 100:.1f}%".replace(".", ",")


def variacao_ml(atual, anterior):
    atual = float(atual or 0)
    anterior = float(anterior or 0)

    if anterior == 0:
        return "-"

    v = ((atual - anterior) / anterior) * 100
    texto = f"{abs(v):.1f}%".replace(".", ",")

    if v > 0:
        return f"▲ {texto}"

    if v < 0:
        return f"▼ {texto}"

    return "0%"


def encontrar_cabecalho(df_raw):
    for i in range(min(40, len(df_raw))):
        linha = df_raw.iloc[i].fillna("").astype(str).str.lower().tolist()

        if any(("id do anúncio" in c) or ("id do anuncio" in c) for c in linha):
            return i

    return None


def carregar_relatorios():
    registros = []

    for arquivo in sorted(PASTA_RELATORIOS.glob("*.xlsx")):
        data_relatorio = extrair_data(arquivo.name)

        if not data_relatorio:
            print(f"Arquivo ignorado, sem data no nome: {arquivo.name}")
            continue

        df_raw = pd.read_excel(arquivo, header=None)
        header_index = encontrar_cabecalho(df_raw)

        if header_index is None:
            print(f"Arquivo ignorado, sem cabeçalho esperado: {arquivo.name}")
            continue

        df = pd.read_excel(arquivo, header=header_index)
        df.columns = [str(c).strip() for c in df.columns]

        col_id = achar_coluna(df, ["ID do anúncio", "ID do anuncio"])
        col_vendas = achar_coluna(df, ["Vendas brutas", "Venda bruta"])
        col_visitas = achar_coluna(df, ["Visitas únicas", "Visitas unicas", "Visitas"])
        col_qtd = achar_coluna(df, ["Quantidade de vendas", "Qtd vendas"])
        col_participacao = achar_coluna(df, ["% de participação", "% de participacao", "participação", "participacao"])
        col_conversao = achar_coluna(df, ["Conversão de vendas", "Conversão", "Conversao"])
        col_qualidade = achar_coluna(df, ["Qualidade do anúncio", "Qualidade do anuncio", "Qualidade"])
        col_experiencia = achar_coluna(df, ["Experiência de compra", "Experiencia de compra", "Experiência", "Experiencia"])

        if col_id is None:
            print(f"Arquivo ignorado, sem coluna ID do anúncio: {arquivo.name}")
            continue

        for _, r in df.iterrows():
            mlb = str(valor_linha(r, col_id, "")).strip()

            if not mlb or mlb.lower() == "nan":
                continue

            registros.append({
                "data_relatorio": data_relatorio,
                "mlb": mlb,
                "vendas": limpar_numero(valor_linha(r, col_vendas, 0)),
                "visitas": limpar_inteiro_ml(valor_linha(r, col_visitas, 0)),
                "quantidade": limpar_inteiro_ml(valor_linha(r, col_qtd, 0)),
                "participacao": limpar_pct(valor_linha(r, col_participacao, 0)),
                "conversao": limpar_pct(valor_linha(r, col_conversao, 0)),
                "qualidade": str(valor_linha(r, col_qualidade, "-")).strip(),
                "experiencia": str(valor_linha(r, col_experiencia, "-")).strip()
            })

    base = pd.DataFrame(registros)

    if base.empty:
        return base

    base = base.sort_values(
        by=["data_relatorio", "mlb", "vendas", "visitas"],
        ascending=[True, True, False, False]
    )

    base = base.drop_duplicates(
        subset=["data_relatorio", "mlb"],
        keep="first"
    )

    return base


def carregar_historico():
    if not ARQUIVO_HISTORICO.exists():
        return pd.DataFrame()

    try:
        with open(ARQUIVO_HISTORICO, "r", encoding="utf-8") as f:
            dados = json.load(f)
    except json.JSONDecodeError:
        print("historico_geral.json vazio ou inválido. Criando histórico novo.")
        return pd.DataFrame()

    if not dados:
        return pd.DataFrame()

    df = pd.DataFrame(dados)

    if "data_relatorio" in df.columns:
        df["data_relatorio"] = pd.to_datetime(df["data_relatorio"]).dt.date

    return df


def salvar_historico(df):
    df_salvar = df.copy()

    if df_salvar.empty:
        return

    df_salvar["data_relatorio"] = df_salvar["data_relatorio"].astype(str)

    with open(ARQUIVO_HISTORICO, "w", encoding="utf-8") as f:
        json.dump(
            df_salvar.to_dict(orient="records"),
            f,
            ensure_ascii=False,
            indent=2
        )


def atualizar_historico(base_nova):
    historico_antigo = carregar_historico()

    if base_nova.empty and historico_antigo.empty:
        return pd.DataFrame()

    if base_nova.empty:
        return historico_antigo

    if historico_antigo.empty:
        historico_final = base_nova.copy()
    else:
        historico_final = pd.concat(
            [historico_antigo, base_nova],
            ignore_index=True
        )

    historico_final = historico_final.sort_values(
        by=["data_relatorio", "mlb", "vendas", "visitas"],
        ascending=[True, True, False, False]
    )

    historico_final = historico_final.drop_duplicates(
        subset=["data_relatorio", "mlb"],
        keep="last"
    )

    salvar_historico(historico_final)

    return historico_final


def aplicar_formatacao_excel(ws):
    azul = "0B57D0"
    branco = "FFFFFF"
    verde = "008000"
    vermelho = "D93025"
    cinza_borda = "D9D9D9"

    fill_header = PatternFill(start_color=azul, end_color=azul, fill_type="solid")
    font_header = Font(bold=True, color=branco)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin", color=cinza_borda),
        right=Side(style="thin", color=cinza_borda),
        top=Side(style="thin", color=cinza_borda),
        bottom=Side(style="thin", color=cinza_borda)
    )

    for cell in ws[1]:
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = thin_border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

            if isinstance(cell.value, str):
                if "▲" in cell.value:
                    cell.font = Font(color=verde, bold=True)
                elif "▼" in cell.value:
                    cell.font = Font(color=vermelho, bold=True)

    larguras = {
        "A": 15,
        "B": 28,
        "C": 14,
        "D": 15,
        "E": 16,
        "F": 34,
        "G": 12,
        "H": 30,
        "I": 14,
        "J": 20,
        "K": 18,
        "L": 22,
        "M": 22
    }

    for col, width in larguras.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 42


def gerar_planilhas():
    mapa = pd.read_excel(ARQUIVO_MAPEAMENTO)
    mapa.columns = [str(c).strip().upper() for c in mapa.columns]

    base_nova = carregar_relatorios()
    base = atualizar_historico(base_nova)

    if base.empty:
        print("Nenhum dado encontrado nos relatórios e nenhum histórico salvo.")
        return

    todas_datas = sorted(base["data_relatorio"].unique(), reverse=True)

    for sku, grupo in mapa.groupby("SKU"):
        linhas = []
        itens_sku = []

        for _, item in grupo.iterrows():
            mlb = str(item["MLB"]).strip()

            valor_produto = limpar_numero(item.get("VALOR PRODUTO", 0))
            valor_produto_txt = br_moeda(valor_produto) if valor_produto > 0 else "-"

            hist = base[base["mlb"] == mlb].set_index("data_relatorio").to_dict("index")

            itens_sku.append({
                "mlb": mlb,
                "valor_produto_txt": valor_produto_txt,
                "hist": hist
            })

        for data_rel in todas_datas:
            for item_sku in itens_sku:
                mlb = item_sku["mlb"]
                valor_produto_txt = item_sku["valor_produto_txt"]
                hist = item_sku["hist"]

                data_analise = data_rel + timedelta(days=1)
                atual = hist.get(data_rel)

                data_ant = data_rel - timedelta(days=1)
                anterior = hist.get(data_ant, {
                    "vendas": 0,
                    "visitas": 0
                })

                if not atual:
                    linhas.append([
                        data_analise.strftime("%d/%m/%Y"),
                        data_rel.strftime("%d/%m/%Y"),
                        mlb,
                        valor_produto_txt,
                        "-",
                        "-",
                        "-",
                        "-",
                        "-",
                        "-",
                        "-",
                        "-",
                        "-"
                    ])
                    continue

                linhas.append([
                    data_analise.strftime("%d/%m/%Y"),
                    data_rel.strftime("%d/%m/%Y"),
                    mlb,
                    valor_produto_txt,
                    br_moeda(atual["vendas"]),
                    variacao_ml(atual["vendas"], anterior["vendas"]),
                    int(atual["visitas"]),
                    variacao_ml(atual["visitas"], anterior["visitas"]),
                    br_pct(atual["conversao"]),
                    int(atual["quantidade"]),
                    br_pct(atual["participacao"]),
                    atual["qualidade"] if atual["qualidade"] else "-",
                    atual["experiencia"] if atual["experiencia"] else "-"
                ])

        df_saida = pd.DataFrame(linhas, columns=COLUNAS_FINAIS)

        arquivo_saida = PASTA_SAIDA / f"SKU_{sku}.xlsx"
        arquivo_json = PASTA_JSON / f"SKU_{sku}.json"

        df_saida.to_json(
            arquivo_json,
            orient="records",
            force_ascii=False,
            indent=2
        )

        with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
            df_saida.to_excel(writer, sheet_name="Analise", index=False)

            ws = writer.book["Analise"]
            aplicar_formatacao_excel(ws)

        print(f"Gerado: {arquivo_saida.name}")

    print("Finalizado.")


if __name__ == "__main__":
    gerar_planilhas()